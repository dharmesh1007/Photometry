import astropy.io.fits
import numpy as np
from Brightest import brightestpixel
from astropy import units as u
from astropy.coordinates import SkyCoord
from photutils import CircularAperture, CircularAnnulus
import matplotlib.pyplot as plt
from astropy.wcs import WCS
from pathlib import Path
from openpyxl import Workbook
from counts import counts
from magnitude import mag_calc

""" 1. GLOBAL VARIABLES ------------------------------------------------------- """

images = False          # Display aperture and annulus overlayed on each star.
ap_rad = 1.6            # Aperture radius
an_in = 8               # Inner annulus radius
an_out = 11             # Outer annulus radius
pixmaprad = 5           # Half width of square grid over which to find brightest pixel
save_to_sprdsht = True  # Save data to spreadsheet

# Right ascension and declination of object and standard stars in degrees.
obj_coords = SkyCoord([327.758225541667, 328.367288, 328.241593, 327.327611,
                    328.021682, 327.602105]*u.deg,
                    [12.6255893888889, 13.350623, 12.451215, 12.824631,
                     12.5168, 13.122719]*u.deg, frame = 'fk5' )

# Create a list containing all the full path names of the FITS files.
p = Path('FITS') # Folder where the FITS files are located.
files = list(p.glob('**/*.fits')) # FITS file names for reference

# Create container for filename, date, MJD, magntitude and magntiude error
data = {"Filename":[], "Date":[], "MJD":np.array([]),\
           "Mag":np.array([]), "Magerr":np.array([])}


""" 2. FIND OBJECT STAR MAGNITUDE FROM EACH OF THE FITS FILES -----------------"""

for f in files[0:]:
    f_name = str(f)     # pass file name into string
    with astropy.io.fits.open(f) as hdul:  # Load FITS file info into object named hdul
        try:
            
            """ 2.1 STORE FILE IMAGE AND HEADER DATA--------------------------- """
            # This data is a 2D numpy data array. The first dimension is the y axis value
            # and the second dimension is the x axis value so to index row 5, column 10
            # you would type nparrayname[5,10].
            imgdta = hdul[0].data # create shortcut to image data
            headerinfo = hdul[0].header # Store header info
            f_name = str(f)  # pass file name into string
            
            
            """ 2.2 ACQUIRE CORRELATION BETWEEN PIXEL AND WORLD COORDINATE AXIS """
            w = WCS(hdul[0]) 
            
  
            """ 2.3 CONVERT SKY COORDINATES OF OBJECTS TO PIXEL COORDINATES ---"""
            pix_val = obj_coords.to_pixel(w) # 2D numpy array ([x coords], [y coords])
            
            
            """ 2.4 REVISE PIXEL COORDINATES TO LOCATION OF BRIGHTEST PIXEL ---"""
            # Returns list of two tuples: [(x1, y1), (x2, y2), ...]
            obj_loc = brightestpixel(imgdta, pix_val[0], pix_val[1], pixmaprad)
            
            
            """ 2.5 APERTURE AND ANNULUS DATA ---------------------------------"""
            # Requires object pixel coordinates and radii of aperture and annulus
            aperture = CircularAperture(obj_loc, r=ap_rad)
            annulus = CircularAnnulus(obj_loc, r_in=an_in, r_out=an_out)
            
            
            """ 2.6 DISPLAY THE APERTURE ON THE IMAGE TO CHECK CENTERING ------"""
            if images == True:
                # DISPLAY INDIVIDUAL TARGETS
                for im_plot in range(0, len(obj_loc)):
                    plt.figure(figsize=(15,15))
                    plt.imshow(imgdta, cmap='nipy_spectral')
                    aperture.plot(color='white', lw=2)
                    annulus.plot(color='red', lw=2)
                    xcent = int(obj_loc[im_plot][0])
                    ycent = int(obj_loc[im_plot][1])
                    plt.xlim(xcent-15, xcent+15)
                    plt.ylim(ycent-15, ycent+15)

                # DISPLAY WHOLE REGION OF AG PEG AND STANDARD STARS
                plt.figure(figsize=(20, 20))
                plt.imshow(imgdta, cmap='nipy_spectral')
                aperture.plot(color='white', lw=2)
                annulus.plot(color='red', lw=2)

            
            """ 2.7 OBJECT MAGNITUDE WITH RESPECT TO EACH STANDARD STAR """
            # Counts of object and standard stars with associated errors.
            star_cts = counts(imgdta, aperture, annulus, 5)
            
            # Magnitudes and associated errors of standard star/stars.
            std_mag = np.array([[9.013, 10.067, 10.07, 10.089, 10.124],\
                                [0.003, 0.001, 0.049, 0.001, 0.052]])
            
            # Calculate magntiude of object with respect to each standard star.
            mags = mag_calc(star_cts[0][0], star_cts[1][0], star_cts[0][1:], \
                            star_cts[1][1:], std_mag[0], std_mag[1])
            
            # Remove a standard star calculation if you wish
            mags = np.delete(mags, 3, 1) # Removing fourth standard star.
            
            
            """ 2.8 CALCULATE AVERAGE MAGNITUDE AND ASSOCIATED ERROR ----------"""
            mean_mag = np.mean(mags[0])
            mean_magerr = (np.sqrt(np.sum(mags[1]**2)))/len(mags[1])
            
            
            """ 2.9 PLACE ALL RELEVANT DATA INTO DICTIONARY ------------------"""
            data["Filename"].append(f_name[-27:])
            data["Date"].append(headerinfo['DATE'])
            data["MJD"] = np.append(data["MJD"], float(headerinfo['MJD']))
            data["Mag"] = np.append(data["Mag"], mean_mag)
            data["Magerr"] = np.append(data["Magerr"], mean_magerr)
            
            
            """ 2.10 PRINT THE FILENAME WHEN A FILE IS PROCESSED --------------"""
            print(f_name)       # print filename to show progress
            
        except:
            print("%s          ERROR!!!" %f_name) # Indicate files that produce errors.
            continue
        

""" 3. SAVE DATA TO AN EXCEL SPREADSHEET -------------------------------------"""

if save_to_sprdsht == True:
    wb = Workbook()     # Create an excel workbook.
    ws = wb.active      # Within the workbook activate a sheet.
    
    # Create headers for the data.
    headers = ['Filename', 'Date', 'MJD', 'Magnitude', 'Magnitude error']
    
    # Pass headers and relevant data into excel spreadsheet.
    for i in range(0, len(headers)):
        ws.cell(row = 1, column = i+1, value = headers[i])
        
    for i in range(0, len(data["Filename"])):
        ws.cell(row = i+2, column = 1, value = data["Filename"][i])
        ws.cell(row = i+2, column = 2, value = data["Date"][i])
        ws.cell(row = i+2, column = 3, value = data["MJD"][i])
        ws.cell(row = i+2, column = 4, value = data["Mag"][i])
        ws.cell(row = i+2, column = 5, value = data["Magerr"][i])
    
    wb.save('codetest.xlsx') # Save file


""" 4. PLOT LIGHT CURVE -------------------------------------------------------"""

plt.figure(figsize=(20, 10)) # Image size
plt.scatter(data["MJD"], data["Mag"], ) # scatter plot values
plt.ylim((max(data["Mag"]), min(data["Mag"]))) # set limits of y axis to reverse order
plt.xlabel("MJD", size=20)
plt.ylabel("Magnitude", size=20)
plt.xticks(fontsize = 20)
plt.yticks(fontsize = 20)
plt.grid()










